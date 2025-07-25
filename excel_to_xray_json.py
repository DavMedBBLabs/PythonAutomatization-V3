import json
import os
import re
import sys

import requests
from colorama import Fore, Style, init
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
init(autoreset=True)


def get_env(var_name):
    value = os.getenv(var_name)
    if not value:
        print(f"{Fore.RED}Error: Variable de entorno '{var_name}' no definida.{Style.RESET_ALL}")
        sys.exit(1)
    return value


def get_auth_token():
    client_id = get_env("CLIENT_ID")
    client_secret = get_env("CLIENT_SECRET")
    auth_url = get_env("AUTH_URL")

    try:
        response = requests.post(
            auth_url,
            json={"client_id": client_id, "client_secret": client_secret},
            timeout=10,
        )
        response.raise_for_status()

        try:
            data = response.json()
        except ValueError:
            print(f"{Fore.RED}Error: respuesta de autenticaci\u00f3n no es JSON v\u00e1lido.{Style.RESET_ALL}")
            sys.exit(1)

        token = None
        if isinstance(data, dict):
            token = data.get("access_token") or data.get("token") or data.get("jwt")
        elif isinstance(data, str):
            token = data

        if not token:
            print(f"{Fore.RED}Error: respuesta de autenticaci\u00f3n sin token.{Style.RESET_ALL}")
            sys.exit(1)

        return token

    except requests.RequestException as exc:
        print(f"{Fore.RED}Error obteniendo token: {exc}{Style.RESET_ALL}")
        sys.exit(1)


def read_excel_data(excel_path, feature_number):
    print(f"Leyendo archivo Excel: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    filename = os.path.splitext(os.path.basename(excel_path))[0]
    numbers = re.findall(r"\d+", filename)
    block_index = 0
    repo_number = numbers[block_index] if numbers else ""
    repo_folder = f"Feature-{feature_number}/HU-{repo_number}"

    tests = []
    prev_test_id = None
    current_test = None

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None or len(row) < 6:
            print(f"{Fore.YELLOW}Advertencia: fila {idx} no v\u00e1lida, omitida.{Style.RESET_ALL}")
            continue
        raw_test_id, summary, description, step, data_col, expected = row[:6]
        if raw_test_id is None:
            print(f"{Fore.YELLOW}Advertencia: fila {idx} sin Test ID, omitida.{Style.RESET_ALL}")
            continue
        try:
            test_id = int(str(raw_test_id).strip())
        except ValueError:
            print(f"{Fore.YELLOW}Advertencia: Test ID inv\u00e1lido en fila {idx}, omitida.{Style.RESET_ALL}")
            continue

        summary = str(summary).strip() if summary else ""
        description = str(description).strip() if description else ""
        step = str(step).strip() if step else ""
        data_col = str(data_col).strip() if data_col else ""
        expected = str(expected).strip() if expected else ""

        if prev_test_id is not None and test_id < prev_test_id:
            block_index += 1
            if block_index < len(numbers):
                repo_number = numbers[block_index]
            repo_folder = f"Feature-{feature_number}/HU-{repo_number}"

        if prev_test_id is None or test_id != prev_test_id:
            current_test = {
                "testtype": "Manual",
                "fields": {
                    "summary": summary,
                    "description": description,
                    "project": {"key": ""},  # placeholder, set later
                    "issuetype": {"name": "Test"},
                },
                "steps": [],
                "xray_test_repository_folder": repo_folder,
                "xray_test_sets": [],
            }
            tests.append(current_test)
        current_test["steps"].append({
            "action": step,
            "data": data_col,
            "result": expected,
        })

        prev_test_id = test_id

    return tests


def generate_xray_json(tests, project_key):
    for test in tests:
        test["fields"]["project"]["key"] = project_key
    return tests


def send_json_to_xray(json_data, auth_token):
    url = get_env("XRAY_IMPORT_URL")
    headers = {
        "Authorization": f"Bearer {auth_token}",
        "Content-Type": "application/json",
    }

    print("Enviando datos a XRay...")
    try:
        response = requests.post(url, json=json_data, headers=headers, timeout=10)
        response.raise_for_status()
        print(f"{Fore.GREEN}Envío exitoso a XRay.{Style.RESET_ALL}")
        if response.text:
            print(response.text)
        return True
    except requests.HTTPError:
        error_text = response.text if hasattr(response, "text") else ""
        print(
            f"{Fore.RED}Error HTTP al enviar a XRay: {response.status_code} {error_text}{Style.RESET_ALL}"
        )
    except requests.RequestException as exc:
        print(f"{Fore.RED}Error de conexión al enviar a XRay: {exc}{Style.RESET_ALL}")
    return False


def process_excel_to_json(file_name, project_key, feature_number):
    base_path = get_env("PATH_EXCEL")
    excel_path = os.path.join(base_path, file_name)
    if not os.path.isfile(excel_path):
        print(f"{Fore.RED}Error: archivo no encontrado: {excel_path}{Style.RESET_ALL}")
        return

    token = get_auth_token()
    print(f"Token obtenido: {token[:5]}...")
    tests = read_excel_data(excel_path, feature_number)
    data = generate_xray_json(tests, project_key)

    json_dir = get_env("PATH_JSON")
    os.makedirs(json_dir, exist_ok=True)
    json_path = os.path.join(json_dir, os.path.splitext(file_name)[0] + ".json")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Archivo JSON guardado en: {json_path}")

    if input("¿Enviar JSON a XRay? (s/n): ").strip().lower() == "s":
        send_json_to_xray(data, token)

    print("Proceso completado.")


def main_menu():
    while True:
        print("\n1. Procesar Excel a JSON")
        print("2. Salir")
        choice = input("Seleccione una opci\u00f3n: ").strip()
        if choice == "1":
            project_key = input("Ingrese la clave del proyecto: ").strip()
            feature_number = input("Ingrese el n\u00famero de Feature: ").strip()
            file_name = input("Ingrese el nombre del archivo Excel: ").strip()
            process_excel_to_json(file_name, project_key, feature_number)
        elif choice == "2":
            break
        else:
            print("Opcion no v\u00e1lida")


if __name__ == "__main__":
    main_menu()